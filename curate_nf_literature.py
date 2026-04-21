"""
NF Guideline Literature Curation Automation
Reads publication titles from Excel, fetches PubMed abstracts,
classifies with Claude, and generates summary counts.
"""

import json
import time
import re
import os
import sys
import io
from pathlib import Path
import requests
import pdfplumber
import openpyxl
try:
    import yaml
except ImportError:
    yaml = None
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic

# Load creds.yaml if present
_creds_file = Path(__file__).parent / "creds.yaml"
if _creds_file.exists() and yaml:
    _creds = yaml.safe_load(_creds_file.read_text())
    if _creds.get("anthropic_api_key"):
        os.environ.setdefault("ANTHROPIC_API_KEY", _creds["anthropic_api_key"])
    if _creds.get("ncbi_api_key"):
        os.environ.setdefault("NCBI_API_KEY", _creds["ncbi_api_key"])

# Load .env file as fallback
_env_file = Path(__file__).parent / ".env"
if _env_file.exists():
    for line in _env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_FILE = "Tables for NF Conference with articles for team review.xlsx"
OUTPUT_FILE = "NF_Guidelines_Curated.xlsx"
CACHE_FILE = "pubmed_cache.json"
RESULTS_FILE = "classification_results.json"
CACHE_FILE_V2 = "pubmed_cache_v2.json"      # PMID-keyed cache for query-based runs
RESULTS_FILE_V2 = "classification_results_v2.json"
SEARCH_PMIDS_FILE = "search_pmids.json"

PUBMED_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
NCBI_API_KEY = os.environ.get("NCBI_API_KEY", "")  # optional but raises rate limit
UNPAYWALL_EMAIL = os.environ.get("UNPAYWALL_EMAIL", "robert.allaway@sagebase.org")
FULLTEXT_MAX_CHARS = 40_000  # ~10k tokens; enough for methods/recommendations sections

def _get_client():
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        sys.exit("ERROR: ANTHROPIC_API_KEY not set. Add it to your environment or create a .env file.")
    return anthropic.Anthropic(api_key=key)

CLIENT: anthropic.Anthropic | None = None  # initialized lazily in main()

# ── Topic taxonomy ───────────────────────────────────────────────────────────
TAXONOMY = {
    "NF1": {
        "manifestations": [
            "Optic pathway glioma",
            "Non-optic pathway glioma",
            "High grade glioma",
            "Low grade glioma",
            "Cutaneous neurofibroma",
            "Diffuse infiltrating neurofibroma",
            "Plexiform neurofibroma",
            "Atypical neurofibroma",
            "ANNUBP",
            "MPNST / triton tumor",
            "Scoliosis",
            "Bone dysplasia",
            "Vascular disease / anomaly",
            "Cognitive",
            "Behavioral",
            "Learning",
        ]
    },
    "NF2-SWN": {
        "manifestations": [
            "Vestibular schwannoma",
            "Meningioma",
            "Peripheral schwannoma",
            "Paraspinal schwannoma",
            "Ependymoma",
            "Hearing loss / Deafness",
            "Facial weakness",
            "Neuropathy",
        ]
    },
    "LZTR1-SWN": {
        "manifestations": [
            "Pain",
            "Intracranial schwannoma",
            "Peripheral schwannoma",
            "Paraspinal schwannoma",
        ]
    },
    "SMARCB1-SWN": {
        "manifestations": [
            "Pain",
            "Peripheral schwannoma",
            "Paraspinal schwannoma",
            "Meningioma",
        ]
    },
}

METHODOLOGIES = [
    "Systematic review & meta-analysis",
    "GRADE (Evidence-to-Decision)",
    "Diagnostic test appraisal",
    "ACMG/AMP variant classification framework",
    "Delphi method (modified)",
    "Nominal Group Technique (NGT)",
    "RAND/UCLA Appropriateness Method",
    "Consensus conference / expert panel",
    "Living guideline model",
    "AGREE II (appraisal tool)",
    "Patient/public & qualitative methods",
]

STRENGTH_LEVELS = ["Strong", "Moderate", "Weak/Conditional", "Expert opinion only", "Not assessable"]


# ── PubMed helpers ───────────────────────────────────────────────────────────
def load_cache():
    if Path(CACHE_FILE).exists():
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def clean_title(raw: str) -> str:
    """Strip '- PubMed' suffix and extra whitespace."""
    return re.sub(r"\s*-\s*PubMed\s*$", "", raw).strip()


def search_pubmed(title: str) -> str | None:
    """Return PMID for a paper title using progressive search strategies."""
    base_params = {"db": "pubmed", "retmax": 3, "retmode": "json"}
    if NCBI_API_KEY:
        base_params["api_key"] = NCBI_API_KEY

    # Strategy 1: key words from title as Title field search (most reliable)
    # Take first 6 meaningful words
    words = [w for w in title.split() if len(w) > 3][:7]
    short_query = " ".join(words) + "[Title]"

    strategies = [
        f'"{title}"[Title]',
        short_query,
        title[:100],
    ]
    try:
        for term in strategies:
            params = {**base_params, "term": term}
            r = requests.get(f"{PUBMED_BASE}/esearch.fcgi", params=params, timeout=15)
            r.raise_for_status()
            ids = r.json()["esearchresult"]["idlist"]
            if ids:
                return ids[0]
        return None
    except Exception as e:
        print(f"  [search error] {e}")
        return None


def fetch_abstract(pmid: str) -> dict:
    """Return {pmid, title, abstract, year, journal, doi, pmc_id, fulltext, fulltext_source} from PubMed."""
    params = {"db": "pubmed", "id": pmid, "retmode": "json"}
    if NCBI_API_KEY:
        params["api_key"] = NCBI_API_KEY
    try:
        r = requests.get(f"{PUBMED_BASE}/esummary.fcgi", params=params, timeout=15)
        r.raise_for_status()
        result = r.json()["result"][pmid]
        year = result.get("pubdate", "")[:4]
        title = result.get("title", "")
        journal = result.get("source", "")
        doi = ""
        pmc_id = ""
        for art in result.get("articleids", []):
            if art["idtype"] == "doi":
                doi = art["value"]
            if art["idtype"] == "pmc":
                pmc_id = art["value"]

        # Abstract via efetch XML
        r2 = requests.get(
            f"{PUBMED_BASE}/efetch.fcgi",
            params={"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"},
            timeout=15,
        )
        abstract = ""
        if r2.status_code == 200:
            parts = re.findall(r"<AbstractText[^>]*>(.*?)</AbstractText>", r2.text, re.DOTALL)
            abstract = " ".join(re.sub(r"<[^>]+>", " ", p).strip() for p in parts)
            # Book articles (e.g. CADTH reports) store title in <BookTitle>
            if not title:
                book_title = re.findall(r"<BookTitle[^>]*>(.*?)</BookTitle>", r2.text, re.DOTALL)
                if book_title:
                    title = re.sub(r"<[^>]+>", "", book_title[0]).strip()

        data = {"pmid": pmid, "title": title, "abstract": abstract, "year": year,
                "journal": journal, "doi": doi, "pmc_id": pmc_id,
                "fulltext": "", "fulltext_source": "abstract_only"}

        # Try full text
        fulltext, source = fetch_fulltext(pmid, pmc_id, doi)
        if fulltext:
            data["fulltext"] = fulltext[:FULLTEXT_MAX_CHARS]
            data["fulltext_source"] = source

        return data
    except Exception as e:
        print(f"  [fetch error pmid={pmid}] {e}")
        return {"pmid": pmid, "title": "", "abstract": "", "year": "", "journal": "",
                "doi": "", "pmc_id": "", "fulltext": "", "fulltext_source": "error"}


def fetch_fulltext(pmid: str, pmc_id: str, doi: str) -> tuple[str, str]:
    """Try PMC XML, then Unpaywall PDF. Returns (text, source_label)."""

    # Strategy 1: PMC full text XML (best quality, structured)
    if pmc_id:
        numeric_id = re.sub(r"[^\d]", "", pmc_id)
        try:
            r = requests.get(
                f"{PUBMED_BASE}/efetch.fcgi",
                params={"db": "pmc", "id": numeric_id, "retmode": "xml", "rettype": "full"},
                timeout=30,
            )
            if r.status_code == 200 and "<body" in r.text.lower():
                # Extract readable text: body sections, strip XML tags
                body_match = re.search(r"<body>(.*?)</body>", r.text, re.DOTALL | re.IGNORECASE)
                raw = body_match.group(1) if body_match else r.text
                text = re.sub(r"<[^>]+>", " ", raw)
                text = re.sub(r"\s+", " ", text).strip()
                if len(text) > 500:
                    print(f"    -> full text via PMC XML ({len(text):,} chars)")
                    return text, "PMC XML"
        except Exception as e:
            print(f"    [PMC XML error] {e}")

    # Strategy 2: Unpaywall open-access PDF
    if doi:
        try:
            r = requests.get(
                f"https://api.unpaywall.org/v2/{doi}",
                params={"email": UNPAYWALL_EMAIL},
                timeout=15,
            )
            if r.status_code == 200:
                data = r.json()
                if data.get("is_oa"):
                    for loc in data.get("oa_locations", []):
                        pdf_url = loc.get("url_for_pdf")
                        if pdf_url:
                            text = download_pdf_text(pdf_url)
                            if text:
                                print(f"    -> full text via Unpaywall PDF ({len(text):,} chars)")
                                return text, f"PDF:{pdf_url}"
        except Exception as e:
            print(f"    [Unpaywall error] {e}")

    return "", ""


def download_pdf_text(url: str) -> str:
    """Download a PDF and extract its text. Returns empty string on failure."""
    try:
        r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200 and b"%PDF" in r.content[:10]:
            with pdfplumber.open(io.BytesIO(r.content)) as pdf:
                pages = [page.extract_text() or "" for page in pdf.pages]
            text = "\n".join(pages)
            text = re.sub(r"\s+", " ", text).strip()
            return text
    except Exception as e:
        print(f"    [PDF error] {e}")
    return ""


# ── Classification prompt ────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a medical literature curator specializing in neurofibromatosis (NF) and schwannomatosis guidelines.
Classify each paper accurately based on its title, abstract, and full text when provided. Return ONLY valid JSON — no markdown, no explanation."""

DOCUMENT_TYPES = [
    "Clinical practice guideline",       # Formal, evidence-based, issued by professional body
    "Consensus statement",               # Expert panel consensus, Delphi, NGT process
    "Evidence-based recommendation",     # Systematic review with explicit recommendations
    "Expert recommendation",             # Expert opinion / narrative recommendation without full guideline process
    "Practice advisory",                 # Informal guidance, commentary with actionable suggestions
    "Not a guideline or recommendation", # Research study, review, case series, clinical trial, etc.
]

def build_user_prompt(title: str, abstract: str, fulltext: str = "", fulltext_source: str = "") -> str:
    method_list = "\n".join(f"- {m}" for m in METHODOLOGIES)
    doc_type_list = "\n".join(f"- {t}" for t in DOCUMENT_TYPES)
    text_section = f"ABSTRACT: {abstract if abstract else '(not available)'}"
    if fulltext:
        text_section += f"\n\nFULL TEXT ({fulltext_source}):\n{fulltext}"
    return f"""Classify this paper for the CTF NF Guideline Steering Committee.

TITLE: {title}
{text_section}

Return JSON with exactly these fields:
{{
  "document_type": "",  // Exactly one of:
{doc_type_list}
  "is_guideline_or_recommendation": true/false,  // true for any type except "Not a guideline or recommendation"
  "conditions": [],  // List any that apply: "NF1", "NF2-SWN", "LZTR1-SWN", "SMARCB1-SWN", "ALL"
  "nf1_manifestations": [],  // From: {json.dumps([m for m in TAXONOMY["NF1"]["manifestations"]])}
  "nf2_swn_manifestations": [],  // From: {json.dumps([m for m in TAXONOMY["NF2-SWN"]["manifestations"]])}
  "lztr1_swn_manifestations": [],  // From: {json.dumps([m for m in TAXONOMY["LZTR1-SWN"]["manifestations"]])}
  "smarcb1_swn_manifestations": [],  // From: {json.dumps([m for m in TAXONOMY["SMARCB1-SWN"]["manifestations"]])}
  "focus_diagnosis": true/false,
  "focus_treatment": true/false,
  "focus_surveillance_management": true/false,
  "focus_genetic_testing": true/false,
  "methodologies": [],  // List any that apply from:
{method_list}
  "strength_of_evidence": "",  // One of: "Strong", "Moderate", "Weak/Conditional", "Expert opinion only", "Not assessable"
  "affiliated_org": "",  // e.g. ERN GENTURIS, AAP, ACMG, CTF, NSGC, EANO, etc.
  "notes": ""  // Brief note if borderline or notable
}}"""


def classify_paper(title: str, abstract: str, fulltext: str = "", fulltext_source: str = "") -> dict:
    """Call Claude to classify a single paper. Returns parsed JSON dict."""
    try:
        msg = CLIENT.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": build_user_prompt(title, abstract, fulltext, fulltext_source)}],
        )
        text_blocks = [b for b in msg.content if b.type == "text"]
        text = text_blocks[0].text.strip()  # type: ignore[attr-defined]
        # Strip any accidental markdown fences
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.MULTILINE).strip()
        return json.loads(text)
    except Exception as e:
        print(f"  [classify error] {e}")
        return {}


# ── Main pipeline ────────────────────────────────────────────────────────────
def read_titles_from_excel() -> list[str]:
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["NF Guidelines to compare"]
    titles = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1]:
            titles.append(str(row[1]).strip())
    return titles


def fetch_all_pubmed(titles: list[str], cache: dict) -> dict:
    """Fetch PubMed data for all titles, using cache."""
    delay = 0.4 if not NCBI_API_KEY else 0.12
    for i, raw_title in enumerate(titles):
        if raw_title in cache:
            continue
        title = clean_title(raw_title)
        print(f"  [{i+1}/{len(titles)}] Searching: {title[:70]}...")
        pmid = search_pubmed(title)
        if pmid:
            data = fetch_abstract(pmid)
        else:
            data = {"pmid": None, "title": title, "abstract": "", "year": "", "journal": ""}
            print(f"    -> PMID not found")
        cache[raw_title] = data
        save_cache(cache)
        time.sleep(delay)
    return cache


def classify_all(titles: list[str], cache: dict, rerun: bool = False) -> dict:
    """Classify all papers with Claude, saving results incrementally.
    Set rerun=True to re-fetch PubMed data and re-classify everything from scratch."""
    if not rerun and Path(RESULTS_FILE).exists():
        with open(RESULTS_FILE) as f:
            results = json.load(f)
    else:
        results = {}

    for i, raw_title in enumerate(titles):
        if raw_title in results and not rerun:
            continue
        pub = cache.get(raw_title, {})
        title = pub.get("title") or clean_title(raw_title)
        abstract = pub.get("abstract", "")
        fulltext = pub.get("fulltext", "")
        fulltext_source = pub.get("fulltext_source", "")
        src_label = f" [{fulltext_source}]" if fulltext else " [abstract only]"
        print(f"  [{i+1}/{len(titles)}] Classifying: {title[:65]}...{src_label}")
        classification = classify_paper(title, abstract, fulltext, fulltext_source)
        results[raw_title] = {"pubmed": pub, "classification": classification}
        with open(RESULTS_FILE, "w") as f:
            json.dump(results, f, indent=2)
        time.sleep(0.1)  # Stay within rate limits
    return results


# ── Excel output ─────────────────────────────────────────────────────────────
def build_output_excel(results: dict):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Individual classifications ──────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "Classified Papers"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    yes_fill = PatternFill("solid", fgColor="C6EFCE")
    no_fill = PatternFill("solid", fgColor="FFCCCC")

    headers = [
        "PMID", "Title", "Year", "Journal",
        "Full Text Source", "Is Guideline/Recommendation", "Document Type", "Conditions",
        "NF1 Manifestations", "NF2-SWN Manifestations",
        "LZTR1-SWN Manifestations", "SMARCB1-SWN Manifestations",
        "Focus: Diagnosis", "Focus: Treatment",
        "Focus: Surveillance/Management", "Focus: Genetic Testing",
        "Methodologies", "Strength of Evidence",
        "Affiliated Org", "Notes",
    ]
    ws_detail.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for raw_title, data in results.items():
        pub = data.get("pubmed", {})
        c = data.get("classification", {})
        is_gl = c.get("is_guideline_or_recommendation", False)
        def gl(val):
            return val if is_gl else ""

        row = [
            pub.get("pmid", ""),
            pub.get("title", "") or raw_title,
            pub.get("year", ""),
            pub.get("journal", ""),
            pub.get("fulltext_source", ""),
            "Y" if is_gl else "N",
            c.get("document_type", ""),
            gl("; ".join(c.get("conditions", []))),
            gl("; ".join(c.get("nf1_manifestations", []))),
            gl("; ".join(c.get("nf2_swn_manifestations", []))),
            gl("; ".join(c.get("lztr1_swn_manifestations", []))),
            gl("; ".join(c.get("smarcb1_swn_manifestations", []))),
            gl("Y" if c.get("focus_diagnosis") else "N"),
            gl("Y" if c.get("focus_treatment") else "N"),
            gl("Y" if c.get("focus_surveillance_management") else "N"),
            gl("Y" if c.get("focus_genetic_testing") else "N"),
            gl("; ".join(c.get("methodologies", []))),
            gl(c.get("strength_of_evidence", "")),
            gl(c.get("affiliated_org", "")),
            gl(c.get("notes", "")),
        ]
        ws_detail.append(row)
        r = ws_detail.max_row

        # Color all Y/N cells: col 6 = Is Guideline, cols 13-16 = focus areas
        yn_cols = [6, 13, 14, 15, 16]
        for col_idx in yn_cols:
            cell = ws_detail.cell(row=r, column=col_idx)
            if cell.value == "Y":
                cell.fill = yes_fill
            elif cell.value == "N":
                cell.fill = no_fill

    # Auto-width
    for col in ws_detail.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_detail.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    # ── Sheet 2: Summary counts ───────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary Counts")

    guideline_papers = [
        d for d in results.values()
        if d.get("classification", {}).get("is_guideline_or_recommendation")
    ]
    total = len(results)
    total_gl = len(guideline_papers)

    def count_condition(condition):
        return sum(
            1 for d in guideline_papers
            if condition in d.get("classification", {}).get("conditions", [])
            or "ALL" in d.get("classification", {}).get("conditions", [])
        )

    def count_manifestation(condition_key, manifestation):
        field = f"{condition_key.lower().replace('-', '_').replace('2', '2')}_manifestations"
        # map condition key to field name
        field_map = {
            "NF1": "nf1_manifestations",
            "NF2-SWN": "nf2_swn_manifestations",
            "LZTR1-SWN": "lztr1_swn_manifestations",
            "SMARCB1-SWN": "smarcb1_swn_manifestations",
        }
        f = field_map[condition_key]
        return sum(
            1 for d in guideline_papers
            if manifestation in d.get("classification", {}).get(f, [])
        )

    def count_focus(focus_key):
        return sum(1 for d in guideline_papers if d.get("classification", {}).get(focus_key))

    def count_method(method):
        return sum(
            1 for d in guideline_papers
            if method in d.get("classification", {}).get("methodologies", [])
        )

    section_fill = PatternFill("solid", fgColor="2E4057")
    section_font = Font(color="FFFFFF", bold=True, size=12)
    subheader_fill = PatternFill("solid", fgColor="5B8DB8")
    subheader_font = Font(color="FFFFFF", bold=True)

    def add_section(title):
        ws_summary.append([title])
        cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        cell.fill = section_fill
        cell.font = section_font
        ws_summary.merge_cells(f"A{ws_summary.max_row}:C{ws_summary.max_row}")

    def add_subheader(title):
        ws_summary.append([title])
        cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        cell.fill = subheader_fill
        cell.font = subheader_font
        ws_summary.merge_cells(f"A{ws_summary.max_row}:C{ws_summary.max_row}")

    def add_row(label, count, note=""):
        ws_summary.append([label, count, note])

    def count_doctype(dtype):
        return sum(1 for d in guideline_papers if d.get("classification", {}).get("document_type") == dtype)

    add_section("NF GUIDELINE LITERATURE SUMMARY")
    ws_summary.append(["Total papers screened", total, ""])
    ws_summary.append(["Total guideline/recommendation papers", total_gl, ""])
    ws_summary.append([""])

    add_subheader("BY DOCUMENT TYPE")
    for dtype in DOCUMENT_TYPES[:-1]:  # exclude "Not a guideline"
        add_row(dtype, count_doctype(dtype))
    ws_summary.append([""])

    add_subheader("BY CONDITION")
    for cond in ["NF1", "NF2-SWN", "LZTR1-SWN", "SMARCB1-SWN"]:
        add_row(cond, count_condition(cond))
    ws_summary.append([""])

    add_subheader("BY FOCUS AREA")
    add_row("Diagnosis", count_focus("focus_diagnosis"))
    add_row("Treatment", count_focus("focus_treatment"))
    add_row("Surveillance / Management", count_focus("focus_surveillance_management"))
    add_row("Genetic Testing and Counseling", count_focus("focus_genetic_testing"))
    ws_summary.append([""])

    add_subheader("NF1 MANIFESTATIONS")
    for m in TAXONOMY["NF1"]["manifestations"]:
        add_row(f"  {m}", count_manifestation("NF1", m))
    ws_summary.append([""])

    add_subheader("NF2-SWN MANIFESTATIONS")
    for m in TAXONOMY["NF2-SWN"]["manifestations"]:
        add_row(f"  {m}", count_manifestation("NF2-SWN", m))
    ws_summary.append([""])

    add_subheader("LZTR1 SCHWANNOMATOSIS MANIFESTATIONS")
    for m in TAXONOMY["LZTR1-SWN"]["manifestations"]:
        add_row(f"  {m}", count_manifestation("LZTR1-SWN", m))
    ws_summary.append([""])

    add_subheader("SMARCB1 SCHWANNOMATOSIS MANIFESTATIONS")
    for m in TAXONOMY["SMARCB1-SWN"]["manifestations"]:
        add_row(f"  {m}", count_manifestation("SMARCB1-SWN", m))
    ws_summary.append([""])

    add_subheader("METHODOLOGY")
    for method in METHODOLOGIES:
        add_row(method, count_method(method))
    ws_summary.append([""])

    add_subheader("STRENGTH OF EVIDENCE")
    for level in STRENGTH_LEVELS:
        n = sum(
            1 for d in guideline_papers
            if d.get("classification", {}).get("strength_of_evidence") == level
        )
        add_row(level, n)

    # Column widths
    ws_summary.column_dimensions["A"].width = 55
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 30

    # ── Sheet 3: Per-year breakdown ───────────────────────────────────────────
    ws_year = wb.create_sheet("By Year")
    ws_year.append(["Year", "Total Papers", "Guideline Papers"])
    year_data: dict[str, dict] = {}
    for raw, data in results.items():
        year = data.get("pubmed", {}).get("year", "Unknown") or "Unknown"
        if year not in year_data:
            year_data[year] = {"total": 0, "guideline": 0}
        year_data[year]["total"] += 1
        if data.get("classification", {}).get("is_guideline_or_recommendation"):
            year_data[year]["guideline"] += 1
    for yr in sorted(year_data.keys()):
        ws_year.append([yr, year_data[yr]["total"], year_data[yr]["guideline"]])

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


# ── PMID-based pipeline (v2 — query-first) ───────────────────────────────────
def fetch_all_by_pmid(pmids: list[str]) -> dict:
    """Fetch PubMed + full text for a list of PMIDs, using PMID-keyed cache."""
    cache = json.loads(Path(CACHE_FILE_V2).read_text()) if Path(CACHE_FILE_V2).exists() else {}
    delay = 0.5 if not NCBI_API_KEY else 0.15
    already = sum(1 for p in pmids if p in cache and cache[p].get("fulltext_source") != "error")
    print(f"   {already}/{len(pmids)} already cached")
    for i, pmid in enumerate(pmids):
        if pmid in cache and cache[pmid].get("fulltext_source") != "error":
            continue
        if i % 50 == 0:
            print(f"  [{i+1}/{len(pmids)}]...", flush=True)
        data = fetch_abstract(pmid)
        cache[pmid] = data
        Path(CACHE_FILE_V2).write_text(json.dumps(cache, indent=2))
        time.sleep(delay)
    return cache


def classify_all_by_pmid(pmids: list[str], cache: dict, rerun: bool = False) -> dict:
    """Classify all papers by PMID, saving incrementally to RESULTS_FILE_V2."""
    results = {}
    if not rerun and Path(RESULTS_FILE_V2).exists():
        results = json.loads(Path(RESULTS_FILE_V2).read_text())

    for i, pmid in enumerate(pmids):
        if pmid in results and not rerun:
            continue
        pub = cache.get(pmid, {})
        title = pub.get("title", f"PMID:{pmid}")
        abstract = pub.get("abstract", "")
        fulltext = pub.get("fulltext", "")
        fulltext_source = pub.get("fulltext_source", "")
        src_label = f"[{fulltext_source}]" if fulltext else "[abstract only]"
        print(f"  [{i+1}/{len(pmids)}] {title[:65]}... {src_label}")
        classification = classify_paper(title, abstract, fulltext, fulltext_source)
        results[pmid] = {"pubmed": pub, "classification": classification}
        Path(RESULTS_FILE_V2).write_text(json.dumps(results, indent=2))
        time.sleep(0.1)
    return results


def build_output_excel_v2(results: dict):
    """Build Excel from PMID-keyed results dict."""
    # Re-key by title for build_output_excel compatibility
    rekeyed = {
        d["pubmed"].get("title") or pmid: d
        for pmid, d in results.items()
    }
    build_output_excel(rekeyed)


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    global CLIENT
    CLIENT = _get_client()

    rerun = "--rerun" in sys.argv
    from_queries = "--from-queries" in sys.argv

    if from_queries:
        # ── Query-based pipeline (v2) ─────────────────────────────────────
        print("=== NF Guideline Literature Curation [FROM PUBMED QUERIES] ===\n")
        if not Path(SEARCH_PMIDS_FILE).exists():
            sys.exit(f"ERROR: {SEARCH_PMIDS_FILE} not found. Run the PubMed query step first.")

        with open(SEARCH_PMIDS_FILE) as f:
            data = json.load(f)
        pmids = data["pmids"]
        counts = data.get("counts", {})
        print(f"1. PMIDs from search queries:")
        for k, v in counts.items():
            print(f"   {k}: {v}")
        print(f"   Total unique: {len(pmids)}\n")

        print("2. Fetching PubMed data (cached)...")
        cache = fetch_all_by_pmid(pmids)
        print(f"   Done. {len(cache)} entries.\n")

        print("3. Classifying with Claude Sonnet...")
        results = classify_all_by_pmid(pmids, cache, rerun=rerun)
        print(f"   Done. {len(results)} classified.\n")

        print("4. Building output Excel...")
        build_output_excel_v2(results)
        result_values = list(results.values())

    else:
        # ── Title-based pipeline (v1, legacy) ────────────────────────────
        if rerun:
            print("=== NF Guideline Literature Curation [FULL RERUN] ===\n")
            for f in [CACHE_FILE, RESULTS_FILE]:
                if Path(f).exists():
                    Path(f).unlink()
                    print(f"   Cleared {f}")
        else:
            print("=== NF Guideline Literature Curation ===\n")
            print("   (Pass --from-queries to use PubMed query results; --rerun to clear cache)\n")

        print("1. Reading titles from Excel...")
        titles = read_titles_from_excel()
        print(f"   Found {len(titles)} publications\n")

        print("2. Fetching PubMed abstracts (cached)...")
        cache = load_cache()
        already_cached = sum(1 for t in titles if t in cache)
        print(f"   {already_cached}/{len(titles)} already cached")
        cache = fetch_all_pubmed(titles, cache)
        print(f"   Done. {len(cache)} entries in cache.\n")

        print("3. Classifying papers with Claude Sonnet...")
        results_dict = classify_all(titles, cache, rerun=rerun)
        print(f"   Done. {len(results_dict)} papers classified.\n")

        print("4. Building output Excel...")
        build_output_excel(results_dict)
        result_values = list(results_dict.values())

    gl_count = sum(1 for d in result_values if d.get("classification", {}).get("is_guideline_or_recommendation"))
    diag = sum(1 for d in result_values if d.get("classification", {}).get("focus_diagnosis"))
    tx = sum(1 for d in result_values if d.get("classification", {}).get("focus_treatment"))
    print(f"\n=== Summary ===")
    print(f"Total screened:                    {len(result_values)}")
    print(f"Confirmed guidelines/recs:         {gl_count}")
    print(f"Focus on diagnosis:                {diag}")
    print(f"Focus on treatment:                {tx}")
    print(f"\nOutput: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
